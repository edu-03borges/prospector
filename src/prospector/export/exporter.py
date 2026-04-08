"""Módulo de exportação para CSV, Excel e JSON."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Sequence

import pandas as pd
from loguru import logger

from prospector.config.settings import DATA_DIR
from prospector.models.lead import Lead


def _leads_to_df(leads: Sequence[Lead]) -> pd.DataFrame:
    """Converte lista de leads em DataFrame pandas."""
    rows = []
    for lead in leads:
        rows.append({
            "id": lead.id,
            "status": lead.status.value,
            "score": lead.score,
            "prioridade": _priority_label(lead.score),
            "nome": lead.name,
            "tipo": lead.studio_type.value,
            "cidade": lead.city,
            "estado": lead.state,
            "telefone": lead.phone,
            "whatsapp": lead.whatsapp_link,
            "email": lead.email or "",
            "email_estimado": "Sim" if lead.email_is_estimated else "Não",
            "email_confiança_%": int((lead.email_confidence or 0) * 100),
            "site": lead.website or "",
            "instagram": lead.social.instagram or "",
            "facebook": lead.social.facebook or "",
            "youtube": lead.social.youtube or "",
            "endereco": lead.address or "",
            "avaliacao_google": lead.rating or "",
            "num_avaliacoes": lead.review_count or 0,
            "google_maps_url": lead.google_maps_url or "",
            "descricao": lead.description or "",
            "notas": lead.notes or "",
            "ultimo_contato": lead.last_contacted_at.isoformat() if lead.last_contacted_at else "",
            "followups": lead.followup_count,
            "fonte": lead.source,
            "criado_em": lead.created_at.isoformat() if lead.created_at else "",
        })
    return pd.DataFrame(rows)


def _priority_label(score: int) -> str:
    if score >= 75:
        return "Alta"
    if score >= 45:
        return "Média"
    return "Baixa"


def export_csv(leads: Sequence[Lead], filename: str | None = None) -> Path:
    """Exporta leads para CSV."""
    df = _leads_to_df(leads)
    name = filename or f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output = DATA_DIR / "exports" / name
    df.to_csv(output, index=False, encoding="utf-8-sig")  # BOM para Excel BR
    logger.success(f"[Export] CSV exportado: {output} ({len(leads)} leads)")
    return output


def export_excel(leads: Sequence[Lead], filename: str | None = None) -> Path:
    """Exporta leads para XLSX com formatação."""
    df = _leads_to_df(leads)
    name = filename or f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output = DATA_DIR / "exports" / name

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        ws = writer.sheets["Leads"]

        # Auto-ajuste de largura das colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)

        # Cabeçalho em negrito/cor
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    logger.success(f"[Export] Excel exportado: {output} ({len(leads)} leads)")
    return output


def export_json(leads: Sequence[Lead], filename: str | None = None) -> Path:
    """Exporta leads para JSON."""
    name = filename or f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    output = DATA_DIR / "exports" / name
    data = [json.loads(lead.model_dump_json()) for lead in leads]
    with open(output, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2, default=str)
    logger.success(f"[Export] JSON exportado: {output} ({len(leads)} leads)")
    return output
